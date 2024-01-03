import os
import openpyxl


class ExcelReaderWriter:
    def __init__(self, filename):
        self.data = {}
        self.excelFilePath = filename
        self.wb = None
        try:
            wb = openpyxl.load_workbook(filename, read_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                headers = [cell.value for cell in sheet[1]]
                self.data[sheet_name] = {}
                for row in sheet.iter_rows(min_row=2):
                    test_case_name = row[0].value
                    if test_case_name not in self.data[sheet_name]:
                        self.data[sheet_name][test_case_name] = {}
                    for col_index, cell in enumerate(row):
                        col_name = headers[col_index]
                        self.data[sheet_name][test_case_name][col_name] = cell.value
            wb.close()
        except FileNotFoundError:
            print(f"Error : File {filename} not found")
        except PermissionError:
            print(f"Error : Permission denied to access file {filename}")
        except Exception as e:
            print(f"Error : {e}")

    def getTestData(self, sheetName, testCaseName, colName):
        if sheetName in self.data and testCaseName in self.data[
                sheetName] and colName in self.data[sheetName][testCaseName]:
            return self.data[sheetName][testCaseName][colName]
        else:
            return f'No data found in Sheetname-{sheetName}, testcase-{testCaseName}, column name-{colName}'

    def setTestData(self, sheetName, testCaseName, colName , orderNumber ):
        self.wb = openpyxl.load_workbook(self.excelFilePath)
        sheet = self.wb[sheetName]
        colIndex = 1
        maxRow = sheet.max_row
        print(maxRow)
        for i in range(1, maxRow + 1):
            print("Entered in Write For loop")
            testName = sheet.cell(i, 1).value
            print(testName)
            if testName == testCaseName:
                while (sheet.cell(row=1, column=colIndex).value != ''):
                    if (colName == sheet.cell(row=1, column=colIndex).value):
                        break
                    colIndex = colIndex + 1
                sheet.cell(i, colIndex).value = orderNumber
                print(sheet.cell(i, colIndex).value)
                break
        print("Come out of For Write loop")
        self.wb.save(self.excelFilePath)
        self.wb.close()

        # Update the data attribute
        if sheetName in self.data and testCaseName in self.data[sheetName]:
            self.data[sheetName][testCaseName][colName] = orderNumber

    def readAllDataInGivenColumn(self, sheetName, colName):
        if sheetName in self.data:
            column_data = []
            for testCaseName in self.data[sheetName]:
                if colName in self.data[sheetName][testCaseName] and self.data[sheetName][testCaseName][colName] != None:
                    column_data.append(self.data[sheetName][testCaseName][colName])
            if column_data:
                return column_data
            else:
                return f'No data found in column name-{colName} in Sheetname-{sheetName}'
        else:
            return f'Sheetname-{sheetName} not found in data'

